from kivy.uix.boxlayout import BoxLayout
from kivy.uix.screenmanager import Screen
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.popup import Popup
from kivy.metrics import dp
from kivy.properties import StringProperty, ListProperty, BooleanProperty
from kivy.graphics import Color, Rectangle
from kivy.utils import get_color_from_hex
from tkinter import Tk, filedialog
from kivy.uix.widget import Widget
from openpyxl import Workbook
from openpyxl.styles import Alignment
from kivy.clock import Clock
from datetime import datetime
import json
import os

class EvaluationCheckContent(BoxLayout):
    def __init__(self, main_layout, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.main_layout = main_layout
        self.padding = dp(20)
        self.all_issues = []
        self.spacing = dp(15)
        self.standard_data = None
        self.user_buttons = {}
        
        # 添加标题
        self.add_widget(Label(
            text="综评数据检查",
            font_size=dp(24),
            size_hint_y=None,
            height=dp(60)
        ))

        # 添加导入标准数据按钮
        self.import_btn = Button(
            text="导入标准数据",
            size_hint=(None, None),
            size=(dp(200), dp(60)),
            pos_hint={'center_x': 0.5},
            background_color=(0.2, 0.6, 0.3, 1),
            color=(1, 1, 1, 1),
            font_size=dp(18)
        )
        self.import_btn.bind(on_press=self.import_json_file)
        self.add_widget(self.import_btn)

        # 添加检查按钮
        self.check_btn = Button(
            text="检查用户数据（自动）",
            size_hint=(None, None),
            size=(dp(200), dp(60)),
            pos_hint={'center_x': 0.5},
            background_color=(0.2, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            font_size=dp(18),
            disabled=True
        )
        self.check_btn.bind(on_press=self.check_user_data)
        self.add_widget(self.check_btn)

        # 添加检查按钮
        self.manual_check_btn = Button(
            text="检查用户数据（手动）",
            size_hint=(None, None),
            size=(dp(200), dp(60)),
            pos_hint={'center_x': 0.5},
            background_color=(0.2, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            font_size=dp(18),
            disabled=True
        )
        self.manual_check_btn.bind(on_press=self.show_manual_check)
        self.add_widget(self.manual_check_btn)

        # 添加导出按钮
        self.generate_btn = Button(
            text="导出结果",
            size_hint=(None, None),
            size=(dp(200), dp(60)),
            pos_hint={'center_x': 0.5},
            background_color=(0.2, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            font_size=dp(18),
            disabled=True
        )
        self.generate_btn.bind(on_press=self.generate_excel)
        self.add_widget(self.generate_btn)

        # 添加状态显示区域
        self.status_label = Label(
            text="请导入JSON格式的标准数据文件",
            font_size=dp(16),
            size_hint_y=None,
            height=dp(40)
        )
        self.add_widget(self.status_label)
        
        # 添加数据显示区域
        self.data_display = ScrollView(size_hint=(1, 1))
        self.data_grid = GridLayout(cols=1, spacing=dp(10), size_hint_y=None)
        self.data_grid.bind(minimum_height=self.data_grid.setter('height'))
        self.data_display.add_widget(self.data_grid)
        self.add_widget(self.data_display)
        
        self.create_user_buttons()

    def create_user_buttons(self):
        """创建所有用户的初始按钮（未检查状态）"""
        self.data_grid.clear_widgets()
        self.user_buttons.clear()

        users = self.main_layout.user_manager.get_all_users()
        if not users:
            self.status_label.text = "没有可检查的用户数据"
            self.status_label.color = (1, 0, 0, 1)
            return

        for username in users:
            user_data = self.main_layout.user_manager.get_user_data(username)
            xm = user_data.get('xm', username)

            user_btn = Button(
                text=f"{xm} ({username}) - 未检查",
                background_color=(0.5, 0.5, 0.5, 1),
                color=(1, 1, 1, 1),
                font_size=dp(16),
                size_hint_y=None,
                height=dp(50),
                disabled=True
            )
            user_btn.bind(on_press=lambda x, u=username: self.show_data_comparison(username = u, type = "auto" ,data={}))
            self.data_grid.add_widget(user_btn)
            self.user_buttons[username] = user_btn
        def update_button_status(self, username, status_text, color):
            if username in self.user_buttons:
                button = self.user_buttons[username]
                button.text = status_text
                button.background_color = color
    def import_json_file(self, instance):
        """导入JSON文件"""  
        try:
            root = Tk()
            root.withdraw()
            file_path = filedialog.askopenfilename(
                title="选择标准数据JSON文件",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )

            if file_path:
                with open(file_path, 'r', encoding='utf-8') as f:
                    self.standard_data = json.load(f)
                if "standard_data" not in self.standard_data:
                    self.status_label.text = "您导入了一个错误的文件！"
                    self.status_label.color = (1, 0, 0, 1)
                else:
                    self.status_label.text = f"成功导入标准数据: {os.path.basename(file_path)}"
                    self.status_label.color = (0, 0.8, 0, 1)
                    self.check_btn.disabled = False
                    self.manual_check_btn.disabled = False
        except Exception as e:
            self.status_label.text = f"导入失败: {str(e)}"
            self.status_label.color = (1, 0, 0, 1)

    def check_user_data(self, instance):
        """检查用户数据并显示结果"""
        self.issues_for_excel = {}
        if not self.standard_data:
            self.status_label.text = "请先导入标准数据"
            self.status_label.color = (1, 0, 0, 1)
            return

        users = self.main_layout.user_manager.get_all_users()
        if not users:
            self.status_label.text = "没有可检查的用户数据"
            self.status_label.color = (1, 0, 0, 1)
            return

        for username in users:
            user_data = self.main_layout.user_manager.get_user_data(username)
            xm = user_data.get('xm', username)

            # 检查是否已有手动标记的问题
            has_manual_issues = False
            for user_issues in self.all_issues:
                if username in user_issues and any(issue["issue_type"] == "manual" for issue in user_issues[username]):
                    has_manual_issues = True
                    break

            # 自动检查
            auto_issues = self.main_layout.compare_user_data(user_data, self.standard_data)
            self.issues_for_excel[xm] = []

            if auto_issues:
                for issue in auto_issues:
                    if issue["issue_type"] == "不匹配":
                        self.issues_for_excel[xm].append((issue["section_name"], issue["key"]))
                    elif issue["issue_type"] == "存在多填/少填的情况":
                        self.issues_for_excel[xm].append((issue["section_name"], "存在多填/少填的情况"))
                    elif issue["issue_type"] == "用户漏填了此项":
                        self.issues_for_excel[xm].append((issue["section_name"], "用户漏填了此项"))

            # 更新按钮状态 - 手动标记的问题优先于自动检查
            if username in self.user_buttons:
                button = self.user_buttons[username]
                if has_manual_issues:
                    button.text = f"{xm} ({username}) - 有问题(存在手动标记的项目)"
                    button.background_color = (0.8, 0.2, 0.2, 1)  # 红色
                elif auto_issues:
                    button.text = f"{xm} ({username}) - 有问题"
                    button.background_color = (0.8, 0.2, 0.2, 1)  # 红色
                else:
                    button.text = f"{xm} ({username}) - 正常"
                    button.background_color = (0.2, 0.6, 0.3, 1)  # 绿色
                button.disabled = False

        self.status_label.text = f"已完成检查，共检查 {len(users)} 个用户"
        self.status_label.color = (0, 0.8, 0, 1)
        self.generate_btn.disabled = False
    def show_data_comparison(self, username, type=None, data=None):
        # 获取用户数据和标准数据
        user_data = self.main_layout.user_manager.get_user_data(username)
        xm = user_data.get('xm', username)

        # 获取所有问题（自动和手动）
        issues = []
        for user_issues in self.all_issues:
            if username in user_issues:
                issues.extend(user_issues[username])

        # 如果是自动检查，添加自动检查的问题
        if type == "auto":
            auto_issues = self.main_layout.compare_user_data(user_data, self.standard_data)
            issues.extend(auto_issues)

        if self.main_layout.manual_check_main_content.user_status.get(username, {}).get("status") == "pass":
            # 用户已通过，移除所有手动issues，只保留自动检查的
            issues = [issue for issue in issues if issue.get("issue_type") != "manual"]
        # 创建弹窗内容布局
        content_layout = BoxLayout(orientation='vertical', padding=dp(20), spacing=dp(15))
        # 标题
        content_layout.add_widget(Label(
            text=f"{xm} 的问题项目",
            font_size=dp(20),
            size_hint_y=None,
            height=dp(40)
        ))
        # 结果滚动区域
        scroll_view = ScrollView(size_hint=(1, 1))
        result_grid = GridLayout(cols=1, spacing=dp(10), size_hint_y=None)
        result_grid.bind(minimum_height=result_grid.setter('height'))
        scroll_view.add_widget(result_grid)
        content_layout.add_widget(scroll_view)
        # 按部分分组整理差异（方便按部分显示）
        section_issues = {}  # 格式：{section_name: [issue1, issue2, ...]}
        for issue in issues:
            section_name = issue["section_name"]
            if section_name not in section_issues:
                section_issues[section_name] = []
            section_issues[section_name].append(issue)

        # 生成比对结果显示内容
        if not issues:
            # 无差异时显示
            no_issues_label = Label(
                text="此用户数据没有任何问题",
                font_size=dp(16),
                color=(0, 0.6, 0, 1),
                size_hint_y=None,
                height=dp(40),
                halign='left'
            )
            result_grid.add_widget(no_issues_label)
        else:
            # 有差异时，按部分显示
            for section_name, section_issue_list in section_issues.items():
                # 部分标题
                section_title = Label(
                    text=f"\n【{section_name}】",
                    font_size=dp(18),
                    color=get_color_from_hex('#0066CC'),
                    size_hint_y=None,
                    height=dp(30),
                    halign='left'
                )
                result_grid.add_widget(section_title)

                # 表格标题行
                table_header = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(40))
                table_header.add_widget(Label(text="信息", font_size=dp(16), bold=True, size_hint_x=0.3))
                table_header.add_widget(Label(text="用户数据", font_size=dp(16), bold=True, size_hint_x=0.25))
                table_header.add_widget(Label(text="标准数据", font_size=dp(16), bold=True, size_hint_x=0.25))
                table_header.add_widget(Label(text="状态", font_size=dp(16), bold=True, size_hint_x=0.2))
                result_grid.add_widget(table_header)

                # 分隔线
                separator = Widget(size_hint_y=None, height=dp(1))
                with separator.canvas:
                    Color(0.7, 0.7, 0.7, 1)
                    Rectangle(pos=separator.pos, size=separator.size)
                result_grid.add_widget(separator)
                # 该部分的所有差异行
                for issue in section_issue_list:
                    issue_type = issue["issue_type"]
                    if issue_type == "不匹配":
                        item_idx = issue["item_idx"]  # 子项索引（第N项）
                        key = issue["key"]  # 字段名
                        user_value = issue["user_value"]
                        std_value = issue["std_value"]  # 标准值
                        row = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(40))
                        row.add_widget(Label(
                            text=f"第{item_idx + 1}项-{key}",
                            font_size=dp(14),
                            size_hint_x=0.3
                        ))
                        row.add_widget(Label(
                            text=str(user_value),
                            font_size=dp(14),
                            color=(1, 0, 0, 1),
                            size_hint_x=0.25
                        ))
                        row.add_widget(Label(
                            text=str(std_value),
                            font_size=dp(14),
                            size_hint_x=0.25
                        ))
                        row.add_widget(Label(
                            text="不符合标准",
                            font_size=dp(14),
                            color=(1, 0, 0, 1),
                            size_hint_x=0.2
                        ))
                        result_grid.add_widget(row)
                    elif issue_type == "存在多填/少填的情况":
                        row = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(40))
                        row.add_widget(Label(
                            text="",
                            font_size=dp(14),
                            size_hint_x=0.3
                        ))
                        row.add_widget(Label(
                            text=f"本项标准数据中只有{issue["std_int"]}个,但此用户却填了{issue["user_int"]}个(存在多填/少填的情况)",
                            font_size=dp(14),
                            color=(1, 0, 0, 1),
                            size_hint_x=0.25
                        ))
                        row.add_widget(Label(
                            text="---",
                            font_size=dp(14),
                            size_hint_x=0.25
                        ))
                        row.add_widget(Label(
                            text="不符合标准",
                            font_size=dp(14),
                            color=(1, 0, 0, 1),
                            size_hint_x=0.2
                        ))
                        result_grid.add_widget(row)
                    elif issue_type == "用户漏填了此项":
                        row = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(40))
                        row.add_widget(Label(
                            text="---",
                            font_size=dp(14),
                            size_hint_x=0.3
                        ))
                        row.add_widget(Label(
                            text="用户漏填了此项",
                            font_size=dp(14),
                            color=(1, 0, 0, 1),
                            size_hint_x=0.25
                        ))
                        row.add_widget(Label(
                            text="---",
                            font_size=dp(14),
                            size_hint_x=0.25
                        ))
                        row.add_widget(Label(
                            text="不符合标准",
                            font_size=dp(14),
                            color=(1, 0, 0, 1),
                            size_hint_x=0.2
                        ))
                        result_grid.add_widget(row)
                    elif issue_type == "manual":
                        row = BoxLayout(orientation='horizontal', size_hint_y=None, height=dp(40))
                        row.add_widget(Label(
                            text="---",
                            font_size=dp(14),
                            size_hint_x=0.3
                        ))
                        row.add_widget(Label(
                            text="---",
                            font_size=dp(14),
                            color=(1, 0, 0, 1),
                            size_hint_x=0.25
                        ))
                        row.add_widget(Label(
                            text=f"错误原因:",
                            font_size=dp(14),
                            size_hint_x=0.25
                        ))
                        row.add_widget(Label(
                            text="不符合标准(手动标记)",
                            font_size=dp(14),
                            color=(1, 0, 0, 1),
                            size_hint_x=0.2
                        ))
                        result_grid.add_widget(row)
        # 底部状态信息
        status_text = f"比对完成，共发现 {len(issues)} 处差异" if issues else "比对完成，未发现差异"
        status_color = (1, 0, 0, 1) if issues else (0, 0.6, 0, 1)
        status_label = Label(
            text=status_text,
            font_size=dp(16),
            color=status_color,
            size_hint_y=None,
            height=dp(30)
        )
        content_layout.add_widget(status_label)

        # 确认按钮
        confirm_button = Button(
            text="关闭",
            size_hint_y=None,
            height=dp(50),
            background_color=(0.2, 0.5, 0.8, 1),
            color=(1, 1, 1, 1),
            font_size=dp(18)
        )
        content_layout.add_widget(confirm_button)

        # 创建并显示弹窗
        popup = Popup(
            title=f"{xm} 的问题项目",
            content=content_layout,
            size_hint=(0.9, 0.9),
            auto_dismiss=False
        )
        confirm_button.bind(on_press=popup.dismiss)
        popup.open()

    def generate_excel(self, instance):
        try:
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            ws.title = "填报情况"

            # 收集所有项目并去重
            all_projects = [
                "任职情况", "自述报告", "相关活动", "日常表现", "个人荣誉",
                "社团活动", "研究性学习", "优势学科", "社会实践", "课外体育修习",
                "体育比赛", "阳光体育出勤", "心理素质展示", "课外艺术修习", "艺术实践活动", "艺术欣赏经历",
            ]

            # 写入表头（项目）
            for col, project in enumerate(all_projects, start=2):
                cell = ws.cell(row=1, column=col, value=project)
                # 设置表头自动换行
                cell.alignment = Alignment(wrapText=True)

            # 合并所有问题（自动检查 + 手动检查）
            all_issues = {}
            
            # 1. 添加自动检查的问题
            for name, projects in self.issues_for_excel.items():
                if name not in all_issues:
                    all_issues[name] = []
                all_issues[name].extend(projects)

            # 2. 添加手动检查的问题
            try:
                # 获取手动检查的数据
                manual_check_main = self.main_layout.manual_check_main_content
                users = self.main_layout.user_manager.get_all_users()
                
                for username in users:
                    user_data = self.main_layout.user_manager.get_user_data(username)
                    xm = user_data.get('xm', username)
                    
                    # 检查该用户是否有手动标记的问题
                    if hasattr(manual_check_main, 'user_status') and username in manual_check_main.user_status:
                        user_status = manual_check_main.user_status[username]
                        
                        # 获取type_status字典
                        type_status = user_status.get('type_status', {})
                        
                        # 遍历所有数据类型索引
                        for data_type_idx, status in type_status.items():
                            if status == "markwrong":
                                # 获取对应的项目名称
                                if hasattr(manual_check_main, 'data_types') and 0 <= int(data_type_idx) < len(manual_check_main.data_types):
                                    project_name = manual_check_main.data_types[int(data_type_idx)]
                                    if xm not in all_issues:
                                        all_issues[xm] = []
                                    # 检查是否已存在相同项目，避免重复
                                    existing_projects = [p[0] for p in all_issues[xm]]
                                    if project_name not in existing_projects:
                                        all_issues[xm].append((project_name, "手动标记错误"))
            except Exception as e:
                print(f"获取手动检查数据时出错: {e}")

            # 写入人名和标记×(具体项目)
            for row, (name, projects) in enumerate(all_issues.items(), start=2):
                # 写入人名
                name_cell = ws.cell(row=row, column=1, value=name)
                name_cell.alignment = Alignment(wrapText=True)

                # 标记×和具体项目
                for project_info in projects:
                    # 解析项目名称和具体描述
                    project_name, project_desc = project_info
                    try:
                        col = all_projects.index(project_name) + 2
                        # 格式化为：×(具体项目)
                        cell = ws.cell(row=row, column=col, value=f"×({project_desc})")
                        # 设置自动换行
                        cell.alignment = Alignment(wrapText=True)
                    except ValueError:
                        # 如果项目名称不在列表中，跳过
                        continue

            # 调整列宽以适应内容
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells)
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

            # 保存Excel文件到当前目录
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"填报情况_{timestamp}.xlsx"
            wb.save(filename)

            # 显示成功提示
            yldwz = self.status_label.text
            self.status_label.text = f"导出成功！文件已保存为：{filename}"
            self.status_label.color = (0, 0.8, 0, 1)  # 绿色
            Clock.schedule_once(lambda dt: setattr(self.status_label, 'text', yldwz), 5)

        except Exception as e:
            # 显示错误提示
            yldwz = self.status_label.text
            self.status_label.text = f"导出失败：{str(e)}"
            self.status_label.color = (1, 0, 0, 1)  # 红色
            Clock.schedule_once(lambda dt: setattr(self.status_label, 'text', yldwz), 5)
    def record_manual_issue(self, username, issue):
        # 先删除该用户该部分的现有手动标记（避免重复）
        section_name = issue.get("section_name")
        if section_name:
            self.remove_manual_issues(username, section_name)
        
        # 查找是否已有该用户的问题记录
        user_found = False
        for user_issues in self.all_issues:
            if username in user_issues:
                user_issues[username].append(issue)
                user_found = True
                break
        if not user_found:
            # 如果没有找到，创建新的记录
            self.all_issues.append({username: [issue]})
        if username in self.user_buttons:
            button = self.user_buttons[username]
            button.text = f"{button.text.split(' - ')[0]} - 有问题"
            button.background_color = (0.8, 0.2, 0.2, 1)  # 红色
    def show_manual_check(self, instance):
        # 切换到手动检查页面
        self.main_layout.show_screen('手动检查')
    def update_user_status(self, username, status_text, color):
        """更新用户状态显示"""
        if username in self.user_buttons:
            button = self.user_buttons[username]
            button.text = status_text
            button.background_color = color
            button.disabled = False
    def remove_manual_issues(self, username, section_name=None):
        """删除指定用户的手动标记issues
        Args:
            username: 用户名
            section_name: 指定部分名称，如果为None则删除所有手动issues
        """
        for user_issues in self.all_issues:
            if username in user_issues:
                if section_name:
                    # 只删除指定部分的手动issues
                    user_issues[username] = [
                        issue for issue in user_issues[username] 
                        if not (issue.get("issue_type") == "manual" and issue.get("section_name") == section_name)
                    ]
                else:
                    # 删除所有手动issues
                    user_issues[username] = [
                        issue for issue in user_issues[username] 
                        if issue.get("issue_type") != "manual"
                    ]
                
                # 如果该用户没有issues了，从列表中移除
                if not user_issues[username]:
                    self.all_issues.remove(user_issues)
                break

    def has_manual_issues(self, username, section_name=None):
        """检查用户是否有手动标记的issues
        Args:
            username: 用户名
            section_name: 指定部分名称，如果为None则检查所有手动issues
        """
        for user_issues in self.all_issues:
            if username in user_issues:
                if section_name:
                    return any(
                        issue.get("issue_type") == "manual" and 
                        issue.get("section_name") == section_name
                        for issue in user_issues[username]
                    )
                else:
                    return any(
                        issue.get("issue_type") == "manual"
                        for issue in user_issues[username]
                    )
        return False
class CheckScreen(Screen):
    def __init__(self, main_layout, **kwargs):
        super().__init__(**kwargs)
        self.add_widget(EvaluationCheckContent(main_layout))